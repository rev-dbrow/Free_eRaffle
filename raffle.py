#!/usr/bin/env python3
#what are you doing here? this isn't for people to actually READ wtf
"""
Click-to-run stages:
1. prepare-- donation credit ticket codes
2. validate-- compare the Google Forms export with official balances
3. draw-- read the approved validation workbook and calculate winners
4. site-- creates presentable website for draw
"""

from __future__ import annotations

import argparse
import hashlib
import json
import math
import re
import secrets
import sys
import unicodedata
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any, Iterable

try:
    import pandas as pd
    from openpyxl import load_workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError as exc:  # pragma: no cover
    raise SystemExit(
        "Required Python packages are missing. Run setup.bat first.\n"
        f"Technical detail: {exc}"
    ) from exc


HEADER_FILL = "1F4E78"
HEADER_FONT = "FFFFFF"
TITLE_FILL = "D9EAF7"
GOOD_FILL = "E2F0D9"
WARNING_FILL = "FFF2CC"
ERROR_FILL = "FCE4D6"
EDITABLE_FILL = "FFF2CC"


@dataclass(frozen=True)
class Paths:
    root: Path
    donations: Path
    volunteers: Path
    pricing: Path
    form_responses: Path
    manual_credits: Path
    preparation: Path
    validation: Path
    results: Path
    website_data: Path
    manual_basket_winners: Path


def now_text() -> str:
    return datetime.now().astimezone().strftime("%Y-%m-%d %H:%M:%S %Z")


def normalize_text(value: Any) -> str:
    if value is None or (isinstance(value, float) and math.isnan(value)):
        return ""
    text = unicodedata.normalize("NFKD", str(value))
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    text = text.replace("’", "'").replace("‘", "'").replace("“", '"').replace("”", '"')
    text = re.sub(r"[^a-zA-Z0-9]+", " ", text.lower())
    return re.sub(r"\s+", " ", text).strip()


def clean_display(value: Any) -> str:
    if value is None or (isinstance(value, float) and math.isnan(value)):
        return ""
    return re.sub(r"\s+", " ", str(value)).strip()


def is_blank(value: Any) -> bool:
    return value is None or (isinstance(value, float) and math.isnan(value)) or not str(value).strip()


def parse_amount(value: Any) -> float | None:
    if is_blank(value):
        return None
    if isinstance(value, (int, float)):
        return round(float(value), 2)
    text = str(value).strip()
    negative = text.startswith("(") and text.endswith(")")
    text = re.sub(r"[^0-9.\-]", "", text)
    if text in {"", "-", ".", "-."}:
        return None
    amount = float(text)
    return round(-amount if negative else amount, 2)


def normalize_code(value: Any) -> str:
    return re.sub(r"\s+", "", clean_display(value)).upper()


def load_config(path: Path) -> dict[str, Any]:
    with path.open("r", encoding="utf-8") as handle:
        config = json.load(handle)
    required = {"event", "files", "donation_columns", "volunteer_columns", "pricing_columns", "form", "validation", "draw"}
    missing = sorted(required - set(config))
    if missing:
        raise ValueError("raffle_config.json is missing section(s): " + ", ".join(missing))
    return config


def resolve_one(root: Path, configured: str, *, required: bool = True) -> Path:
    candidate = root / configured
    if any(char in configured for char in "*?["):
        matches = sorted(root.glob(configured), key=lambda p: p.stat().st_mtime, reverse=True)
        if not matches:
            if required:
                raise FileNotFoundError(f"No file matched: {configured}")
            return candidate
        if len(matches) > 1:
            names = "\n  - ".join(str(p.relative_to(root)) for p in matches)
            raise RuntimeError(
                f"More than one file matched '{configured}'. Keep only the current export:\n  - {names}"
            )
        return matches[0]
    if required and not candidate.exists():
        raise FileNotFoundError(f"File not found: {candidate}")
    return candidate


def build_paths(config_path: Path, config: dict[str, Any], command: str) -> Paths:
    root = config_path.parent.resolve()
    files = config["files"]
    output_dir = root / files.get("output_directory", ".")
    output_dir.mkdir(parents=True, exist_ok=True)

    return Paths(
        root=root,
        donations=(
            resolve_one(root, files["donations"], required=True)
            if command == "prepare"
            else root / files["donations"]
        ),
        volunteers=(
            resolve_one(root, files["volunteers"], required=True)
            if command == "prepare"
            else root / files["volunteers"]
        ),
        pricing=(
            resolve_one(root, files["ticket_pricing"], required=True)
            if command == "prepare"
            else root / files["ticket_pricing"]
        ),
        form_responses=(
            resolve_one(root, files["form_responses"], required=True)
            if command == "validate"
            else root / files["form_responses"]
        ),
        manual_credits=root / files.get("manual_credits", "manual_credits.xlsx"),
        preparation=output_dir / files.get("preparation_workbook", "raffle_preparation.xlsx"),
        validation=output_dir / files.get("validation_workbook", "raffle_validation.xlsx"),
        results=output_dir / files.get("results_workbook", "raffle_results.xlsx"),
        website_data=output_dir / files.get("website_data", "draw_data.js"),
        manual_basket_winners=output_dir / files.get("manual_basket_winners", "manual_basket_winners.xlsx"),
    )


def read_table(path: Path, *, sheet_name: str | int = 0) -> pd.DataFrame:
    suffix = path.suffix.lower()
    if suffix == ".csv":
        frame = pd.read_csv(path, dtype=object, encoding="utf-8-sig")
    elif suffix in {".xlsx", ".xlsm"}:
        frame = pd.read_excel(path, sheet_name=sheet_name, dtype=object)
    else:
        raise ValueError(f"Unsupported file type: {path.suffix}")
    frame.columns = [clean_display(col).lstrip("\ufeff") for col in frame.columns]
    return frame


def require_columns(frame: pd.DataFrame, columns: Iterable[str], label: str) -> None:
    missing = [col for col in columns if col and col not in frame.columns]
    if missing:
        raise ValueError(f"{label} is missing column(s): {', '.join(missing)}")


def first_existing_column(frame: pd.DataFrame, candidates: Iterable[str]) -> str | None:
    normalized = {normalize_text(col): col for col in frame.columns}
    for candidate in candidates:
        if candidate in frame.columns:
            return candidate
        found = normalized.get(normalize_text(candidate))
        if found:
            return found
    return None


def donation_id(row: pd.Series, columns: dict[str, str]) -> str:
    values = [
        clean_display(row.get(columns.get("date", ""), "")),
        clean_display(row.get(columns.get("donor_email", ""), "")),
        clean_display(row.get(columns.get("donor_first", ""), "")),
        clean_display(row.get(columns.get("donor_last", ""), "")),
        clean_display(row.get(columns.get("amount", ""), "")),
        clean_display(row.get(columns.get("comment", ""), "")),
    ]
    return "DON-" + hashlib.sha256("|".join(values).encode("utf-8")).hexdigest()[:10].upper()


def build_volunteer_aliases(
    volunteers: pd.DataFrame, columns: dict[str, str]
) -> tuple[dict[str, set[str]], set[str]]:
    require_columns(volunteers, [columns["first"], columns["last"]], "Volunteer list")
    aliases: dict[str, set[str]] = {}
    # aliases whose roster last name is literally a single letter; these may match
    # a longer last name in a comment (volunteer "Michelle S" <- comment "Michelle Shonka")
    expandable: set[str] = set()

    def add(alias: str, official: str, *, expand: bool = False) -> None:
        key = normalize_text(alias)
        if len(key) >= 2:
            aliases.setdefault(key, set()).add(official)
            if expand:
                expandable.add(key)

    for _, row in volunteers.iterrows():
        first = clean_display(row.get(columns["first"]))
        last = clean_display(row.get(columns["last"]))
        if not first:
            continue
        nickname_col = columns.get("nickname", "")
        nickname = clean_display(row.get(nickname_col)) if nickname_col else ""

        if not last:
            # volunteer listed with a first name only; the exact name is the only
            # usable alias, so allow it despite the first-name-alone risk
            official = first
            add(official, official)
            if nickname:
                add(nickname, official)
        else:
            truncated = len(re.sub(r"[^A-Za-z]", "", last)) == 1
            official = f"{first} {last}"
            add(official, official, expand=truncated)
            add(f"{first} {last[:1]}", official, expand=truncated)
            if nickname:
                add(f"{nickname} {last}", official, expand=truncated)
                add(f"{nickname} {last[:1]}", official, expand=truncated)

        alias_col = columns.get("aliases", "")
        if alias_col and alias_col in volunteers.columns:
            for alias in re.split(r"[;|]", clean_display(row.get(alias_col))):
                if alias.strip():
                    add(alias.strip(), official)
    return aliases, expandable


def match_volunteer(comment: str, aliases: dict[str, set[str]], expandable: set[str] = frozenset()) -> list[str]:
    text = normalize_text(comment)
    if not text:
        return []
    matched: set[str] = set()
    for alias in sorted(aliases, key=lambda item: (-len(item), item)):
        # an expandable alias ends in a truncated roster last name, so the final
        # letter may continue in the comment ("karen a" matches "karen agarcia")
        tail = r"[a-z]*" if alias in expandable else r"(?![a-zA-Z0-9])"
        if re.search(rf"(?<![a-zA-Z0-9]){re.escape(alias)}{tail}", text):
            matched.update(aliases[alias])
    return sorted(matched)


def load_pricing(path: Path, columns: dict[str, str]) -> pd.DataFrame:
    frame = read_table(path)
    require_columns(frame, [columns["donation"], columns["tickets"]], "Ticket pricing")
    rows: list[dict[str, float]] = []
    for _, row in frame.iterrows():
        donation = parse_amount(row.get(columns["donation"]))
        tickets = parse_amount(row.get(columns["tickets"]))
        if donation is None or tickets is None:
            continue
        if donation < 0 or tickets < 0:
            raise ValueError("Ticket pricing cannot contain negative values.")
        rows.append({"Donation": donation, "Tickets": tickets})
    if not rows:
        raise ValueError("The ticket-pricing workbook contains no usable rows.")
    pricing = pd.DataFrame(rows).drop_duplicates("Donation", keep="last").sort_values("Donation")
    return pricing.reset_index(drop=True)


def round_tickets(value: float, mode: str, increment: int) -> int:
    increment = max(1, int(increment))
    scaled = value / increment
    if mode == "floor":
        rounded = math.floor(scaled)
    elif mode == "ceiling":
        rounded = math.ceil(scaled)
    else:
        rounded = math.floor(scaled + 0.5)
    return max(0, int(rounded * increment))


def estimate_tickets(amount: float, pricing: pd.DataFrame, config: dict[str, Any]) -> tuple[int, float, str, str]:
    points = [(float(row.Donation), float(row.Tickets)) for row in pricing.itertuples(index=False)]
    exact = next((tickets for donation, tickets in points if abs(donation - amount) < 0.000001), None)
    if exact is not None:
        raw = exact
        method = "EXACT MATRIX VALUE"
        basis = f"${amount:.2f} = {exact:g} tickets"
    else:
        augmented = points[:]
        if points[0][0] > 0:
            augmented.insert(0, (0.0, 0.0))

        lower: tuple[float, float]
        upper: tuple[float, float]
        if amount < augmented[0][0]:
            lower, upper = (0.0, 0.0), augmented[0]
            method = "BELOW MATRIX RANGE"
        elif amount > augmented[-1][0]:
            if len(augmented) >= 2:
                lower, upper = augmented[-2], augmented[-1]
            else:
                lower, upper = (0.0, 0.0), augmented[-1]
            method = "ABOVE MATRIX RANGE EXTRAPOLATION"
        else:
            lower, upper = augmented[0], augmented[-1]
            method = "PIECEWISE LINEAR INTERPOLATION"
            for left, right in zip(augmented, augmented[1:]):
                if left[0] <= amount <= right[0]:
                    lower, upper = left, right
                    break

        if abs(upper[0] - lower[0]) < 0.000001:
            raw = upper[1]
        else:
            raw = lower[1] + (amount - lower[0]) * (upper[1] - lower[1]) / (upper[0] - lower[0])
        basis = f"(${lower[0]:.2f}, {lower[1]:g}) to (${upper[0]:.2f}, {upper[1]:g})"

    mode = str(config.get("rounding_mode", "nearest")).lower()
    increment = int(config.get("round_to_increment", 1))
    return round_tickets(raw, mode, increment), raw, method, basis


def load_manual_credits(path: Path) -> pd.DataFrame:
    if not path.exists():
        return pd.DataFrame(columns=["Donation ID", "Manual Credit Name", "Manual Ticket Override", "Organizer Notes"])
    frame = read_table(path, sheet_name="Manual Credits")
    required = ["Donation ID", "Manual Credit Name", "Manual Ticket Override", "Organizer Notes"]
    for col in required:
        if col not in frame.columns:
            frame[col] = ""
    return frame[required]


def previous_basket_winners(path: Path) -> dict[str, str]:
    if not path.exists():
        return {}
    try:
        frame = read_table(path, sheet_name="Manual Basket Winners")
    except Exception:
        return {}
    if not {"Basket", "Manual Winner (optional)"}.issubset(frame.columns):
        return {}
    return {
        clean_display(row["Basket"]): clean_display(row["Manual Winner (optional)"])
        for _, row in frame.iterrows()
        if clean_display(row["Basket"]) and clean_display(row["Manual Winner (optional)"])
    }


def previous_codes(preparation_path: Path) -> dict[str, str]:
    if not preparation_path.exists():
        return {}
    try:
        frame = read_table(preparation_path, sheet_name="Ticket Balances")
    except Exception:
        return {}
    if not {"Participant Name", "Ticket Code"}.issubset(frame.columns):
        return {}
    return {
        clean_display(row["Participant Name"]): normalize_code(row["Ticket Code"])
        for _, row in frame.iterrows()
        if clean_display(row["Participant Name"]) and normalize_code(row["Ticket Code"])
    }


def assign_codes(names: Iterable[str], existing: dict[str, str]) -> dict[str, str]:
    names = sorted(
        set(clean_display(name) for name in names if clean_display(name)),
        key=lambda name: (normalize_text(name), name),
    )
    used = {code for code in existing.values() if code}
    output: dict[str, str] = {}
    for name in names:
        if name in existing and existing[name] not in output.values():
            output[name] = existing[name]
            continue
        letters = re.sub(r"[^A-Z]", "", name.upper())
        prefix = (letters[:3] or "TKT").ljust(3, "X")
        number = 1
        while f"{prefix}-{number:03d}" in used:
            number += 1
        code = f"{prefix}-{number:03d}"
        used.add(code)
        output[name] = code
    return output


def dataframe_rows(frame: pd.DataFrame) -> list[list[Any]]:
    return [frame.columns.tolist()] + frame.where(pd.notna(frame), None).values.tolist()


def style_workbook(path: Path, *, status_sheets: Iterable[str] = (), editable_columns: dict[str, list[str]] | None = None) -> None:
    workbook = load_workbook(path)
    editable_columns = editable_columns or {}
    for sheet in workbook.worksheets:
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        for cell in sheet[1]:
            cell.fill = PatternFill("solid", fgColor=HEADER_FILL)
            cell.font = Font(color=HEADER_FONT, bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        sheet.row_dimensions[1].height = 30

        header_map = {clean_display(cell.value): cell.column for cell in sheet[1] if cell.value is not None}
        for header in editable_columns.get(sheet.title, []):
            col_idx = header_map.get(header)
            if col_idx:
                for row in range(2, sheet.max_row + 1):
                    sheet.cell(row, col_idx).fill = PatternFill("solid", fgColor=EDITABLE_FILL)

        if sheet.title in status_sheets:
            good_values = {"EXACT", "YES", "MATCH", "AUTO MATCHED", "REVIEWED", "LATEST RESPONSE USED"}
            warn_values = {
                "UNDER", "WARNING", "NAME DIFFERS", "MISSING NAME",
                "NO VOLUNTEER MATCH - INCLUDED", "OLDER DUPLICATE - NOT USED",
                "NO ACTIVE FORM RESPONSE", "NOT CHECKED",
            }
            for header, col_idx in header_map.items():
                if "Status" in header or header in {"Ready for Draw", "Name Check"}:
                    for row in range(2, sheet.max_row + 1):
                        cell = sheet.cell(row, col_idx)
                        value = clean_display(cell.value).upper()
                        if value in good_values:
                            cell.fill = PatternFill("solid", fgColor=GOOD_FILL)
                        elif value in warn_values or "REVIEW" in value:
                            cell.fill = PatternFill("solid", fgColor=WARNING_FILL)
                        elif value and value not in {"NO RESPONSE"}:
                            cell.fill = PatternFill("solid", fgColor=ERROR_FILL)

        for column_cells in sheet.columns:
            letter = get_column_letter(column_cells[0].column)
            max_length = max(len(clean_display(cell.value)) for cell in column_cells[: min(len(column_cells), 250)])
            width = min(max(max_length + 2, 10), 42)
            sheet.column_dimensions[letter].width = width
        for row in sheet.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
    workbook.save(path)


def write_workbook(path: Path, sheets: list[tuple[str, pd.DataFrame]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        for name, frame in sheets:
            frame.to_excel(writer, sheet_name=name, index=False)


def prepare(config: dict[str, Any], paths: Paths) -> int:
    donations = read_table(paths.donations)
    volunteers = read_table(paths.volunteers)
    pricing = load_pricing(paths.pricing, config["pricing_columns"])
    columns = config["donation_columns"]
    require_columns(
        donations,
        [columns["amount"], columns["donor_first"], columns["donor_last"], columns.get("comment", "")],
        "Donation export",
    )

    aliases, expandable = build_volunteer_aliases(volunteers, config["volunteer_columns"])

    # surface roster quirks the organizer should know about
    vol_cols = config["volunteer_columns"]
    single_name = [
        clean_display(r.get(vol_cols["first"]))
        for _, r in volunteers.iterrows()
        if clean_display(r.get(vol_cols["first"])) and not clean_display(r.get(vol_cols["last"]))
    ]
    if single_name:
        print(f"Volunteers listed with a first name only (matched on that exact name, higher mix-up risk): {', '.join(single_name)}")
    recognized = {vol_cols.get(k, "") for k in ("first", "last", "nickname", "aliases")}
    stray = [
        col for col in volunteers.columns
        if col not in recognized and volunteers[col].map(clean_display).astype(bool).any()
    ]
    if stray:
        print(f"Ignoring extra volunteer-list column(s) that contain data: {', '.join(stray)}")
        print("Names typed in those columns are NOT used for matching. Move them into the First/Last columns if they are volunteers.")

    manual = load_manual_credits(paths.manual_credits)
    manual_by_id = {
        clean_display(row["Donation ID"]): row
        for _, row in manual.iterrows()
        if clean_display(row.get("Donation ID"))
    }

    detail_rows: list[dict[str, Any]] = []
    manual_rows: list[dict[str, Any]] = []
    ticket_config = config.get("ticket_estimation", {})

    for source_index, row in donations.iterrows():
        identifier = donation_id(row, columns)
        amount = parse_amount(row.get(columns["amount"]))
        if amount is None or amount < 0:
            raise ValueError(f"Donation row {source_index + 2} has an invalid amount: {row.get(columns['amount'])}")

        donor_first = clean_display(row.get(columns["donor_first"]))
        donor_last = clean_display(row.get(columns["donor_last"]))
        donor_name = clean_display(f"{donor_first} {donor_last}")
        donor_email = clean_display(row.get(columns.get("donor_email", "")))
        comment = clean_display(row.get(columns.get("comment", "")))
        matches = match_volunteer(comment, aliases, expandable)
        donor_matches = match_volunteer(donor_name, aliases, expandable) if donor_name else []
        manual_row = manual_by_id.get(identifier)
        manual_credit = clean_display(manual_row.get("Manual Credit Name")) if manual_row is not None else ""
        manual_ticket_value = manual_row.get("Manual Ticket Override") if manual_row is not None else None
        organizer_notes = clean_display(manual_row.get("Organizer Notes")) if manual_row is not None else ""

        if manual_credit:
            credited_name = manual_credit
            credit_method = "MANUAL CREDIT"
            credit_status = "REVIEWED"
        elif len(matches) == 1:
            credited_name = matches[0]
            credit_method = "VOLUNTEER NAME IN COMMENT"
            credit_status = "AUTO MATCHED"
        elif len(matches) > 1:
            credited_name = donor_name or f"Unmatched Donation {identifier}"
            credit_method = "DONOR NAME FALLBACK"
            credit_status = "MULTIPLE VOLUNTEERS FOUND - REVIEW"
        elif len(donor_matches) == 1:
            # the comment named nobody, but the donor is a listed volunteer
            # (self-donation); credit the official roster spelling of their name
            credited_name = donor_matches[0]
            credit_method = "DONOR IS VOLUNTEER"
            credit_status = "AUTO MATCHED"
        else:
            credited_name = donor_name or f"Unmatched Donation {identifier}"
            credit_method = "DONOR NAME"
            credit_status = "NO VOLUNTEER MATCH - INCLUDED"

        source_ticket_col = columns.get("tickets", "")
        source_ticket_value = row.get(source_ticket_col) if source_ticket_col and source_ticket_col in donations.columns else None
        manual_ticket = parse_amount(manual_ticket_value)
        source_ticket = parse_amount(source_ticket_value)

        if manual_ticket is not None:
            if manual_ticket < 0 or not float(manual_ticket).is_integer():
                raise ValueError(f"Manual ticket override for {identifier} must be a nonnegative whole number.")
            tickets = int(manual_ticket)
            raw_estimate = float(tickets)
            ticket_method = "MANUAL TICKET OVERRIDE"
            ticket_basis = "manual_credits.xlsx"
        elif source_ticket is not None:
            if source_ticket < 0 or not float(source_ticket).is_integer():
                raise ValueError(f"Source ticket value for {identifier} must be a nonnegative whole number.")
            tickets = int(source_ticket)
            raw_estimate = float(tickets)
            ticket_method = "SOURCE TICKET VALUE"
            ticket_basis = source_ticket_col
        else:
            tickets, raw_estimate, ticket_method, ticket_basis = estimate_tickets(amount, pricing, ticket_config)

        detail_rows.append(
            {
                "Donation ID": identifier,
                "Source Row": source_index + 2,
                "Donation Date": clean_display(row.get(columns.get("date", ""))),
                "Donor Name": donor_name,
                "Donor Email": donor_email,
                "Donation Amount": amount,
                "Donation Comment": comment,
                "Volunteer Matches Found": "; ".join(matches),
                "Donation Credit": credited_name,
                "Credit Method": credit_method,
                "Credit Status": credit_status,
                "Raw Ticket Estimate": raw_estimate,
                "Final Tickets": tickets,
                "Ticket Method": ticket_method,
                "Ticket Basis": ticket_basis,
                "Organizer Notes": organizer_notes,
            }
        )
        manual_rows.append(
            {
                "Donation ID": identifier,
                "Donation Date": clean_display(row.get(columns.get("date", ""))),
                "Donor Name": donor_name,
                "Donation Amount": amount,
                "Donation Comment": comment,
                "Automatic Credit": credited_name,
                "Automatic Credit Status": credit_status,
                "Manual Credit Name": manual_credit,
                "Manual Ticket Override": manual_ticket_value if not is_blank(manual_ticket_value) else "",
                "Organizer Notes": organizer_notes,
            }
        )

    detail = pd.DataFrame(detail_rows)
    # merge credits that differ only in case/spacing ("michelle Shonka" vs "Michelle Shonka"),
    # preferring the official roster spelling when one exists
    official_names = {name for names in aliases.values() for name in names}
    official_by_key = {normalize_text(name): name for name in official_names}
    canonical: dict[str, str] = {}
    for credit in detail["Donation Credit"]:
        key = normalize_text(credit)
        if key not in canonical:
            canonical[key] = official_by_key.get(key, credit)
    detail["Donation Credit"] = detail["Donation Credit"].map(lambda credit: canonical[normalize_text(credit)])
    balances = (
        detail.groupby("Donation Credit", as_index=False)
        .agg(**{"Total Donated": ("Donation Amount", "sum"), "Official Tickets": ("Final Tickets", "sum"), "Donation Count": ("Donation ID", "count")})
        .rename(columns={"Donation Credit": "Participant Name"})
    )
    balances["Total Donated"] = balances["Total Donated"].round(2)
    codes = assign_codes(balances["Participant Name"], previous_codes(paths.preparation))
    balances["Ticket Code"] = balances["Participant Name"].map(codes)
    form_link = clean_display(config["event"].get("form_link", ""))
    balances["Invite Text"] = balances.apply(
        lambda row: (
            f"You have {int(row['Official Tickets'])} raffle tickets. Your ticket code is {row['Ticket Code']}."
            + (f" Submit your choices here: {form_link}" if form_link else "")
        ),
        axis=1,
    )
    balances = balances[["Participant Name", "Total Donated", "Official Tickets", "Ticket Code", "Donation Count", "Invite Text"]].sort_values("Participant Name")

    summary = pd.DataFrame(
        [
            ["Event", config["event"].get("name", "Clinic Fundraising Raffle")],
            ["Event Date", config["event"].get("date", "")],
            ["Prepared At", now_text()],
            ["Donation Export", str(paths.donations.relative_to(paths.root))],
            ["Donation Rows", len(detail)],
            ["Participants", len(balances)],
            ["Total Donations", float(detail["Donation Amount"].sum())],
            ["Total Official Tickets", int(detail["Final Tickets"].sum())],
            ["Comment Matches Needing Review", int(detail["Credit Status"].str.contains("REVIEW", na=False).sum())],
            ["Next Step", "Review manual_credits.xlsx if needed, rerun prepare, then build the Google Form."],
        ],
        columns=["Item", "Value"],
    )

    manual_frame = pd.DataFrame(manual_rows)
    write_workbook(paths.manual_credits, [("Manual Credits", manual_frame)])
    style_workbook(
        paths.manual_credits,
        status_sheets=["Manual Credits"],
        editable_columns={"Manual Credits": ["Manual Credit Name", "Manual Ticket Override", "Organizer Notes"]},
    )

    pricing_out = pricing.copy()
    pricing_out["Tickets per Dollar"] = pricing_out["Tickets"] / pricing_out["Donation"]
    write_workbook(
        paths.preparation,
        [
            ("Summary", summary),
            ("Ticket Balances", balances),
            ("Donation Detail", detail),
            ("Pricing Matrix", pricing_out),
        ],
    )
    style_workbook(paths.preparation, status_sheets=["Donation Detail"])

    print(f"Preparation complete: {paths.preparation}")
    print(f"Manual-credit review file: {paths.manual_credits}")
    print(f"Participants: {len(balances):,} | Official tickets: {int(detail['Final Tickets'].sum()):,}")
    return 0


def find_basket_columns(
    frame: pd.DataFrame, prefix: str, *, reserved: Iterable[str | None], ignore: Iterable[str]
) -> tuple[list[str], bool]:
    normalized_prefix = normalize_text(prefix) if prefix else ""
    prefixed = [col for col in frame.columns if normalized_prefix and normalize_text(col).startswith(normalized_prefix)]
    if prefixed:
        return prefixed, False

    # Fallback: no header begins with the configured prefix, so treat every column
    # that is not a recognized system column (timestamp/name/code) or an ignored
    # column as a basket, using the header text itself as the basket name.
    excluded = {normalize_text(col) for col in reserved if col} | {normalize_text(col) for col in ignore}
    fallback = [col for col in frame.columns if normalize_text(col) not in excluded]
    if not fallback:
        raise ValueError(
            "No basket columns were found in the Google Forms export. "
            "Every column matched the ticket code/name/timestamp columns or an ignore_columns entry."
        )
    return fallback, True


def basket_name(header: str, prefix: str) -> str:
    pattern = re.compile(rf"^\s*{re.escape(prefix)}\s*", flags=re.IGNORECASE)
    name = pattern.sub("", clean_display(header)).strip(" :-")
    return name or clean_display(header)


def parse_ticket_cell(value: Any) -> tuple[int, str]:
    if is_blank(value):
        return 0, ""
    try:
        number = float(str(value).replace(",", "").strip())
    except ValueError:
        return 0, f"NONNUMERIC VALUE: {value}"
    if not math.isfinite(number):
        return 0, f"INVALID VALUE: {value}"
    if number < 0:
        return 0, f"NEGATIVE VALUE: {value}"
    if not number.is_integer():
        return 0, f"FRACTIONAL VALUE: {value}"
    return int(number), ""


def validate(config: dict[str, Any], paths: Paths) -> int:
    if not paths.preparation.exists():
        raise FileNotFoundError("raffle_preparation.xlsx does not exist. Run run_prepare.bat first.")
    if paths.validation.exists():
        raise FileExistsError(
            "raffle_validation.xlsx already exists. Delete it before rerunning validation so there is no confusion about which review is current."
        )

    balances = read_table(paths.preparation, sheet_name="Ticket Balances")
    require_columns(balances, ["Participant Name", "Official Tickets", "Ticket Code"], "Ticket Balances sheet")
    balances["Ticket Code"] = balances["Ticket Code"].map(normalize_code)
    balance_by_code = balances.set_index("Ticket Code").to_dict("index")

    responses = read_table(paths.form_responses)
    form_config = config["form"]
    timestamp_col = first_existing_column(responses, form_config.get("timestamp_columns", ["Timestamp", "Completion time"]))
    name_col = first_existing_column(responses, form_config.get("name_columns", ["Your name", "Name"]))
    code_col = first_existing_column(responses, form_config.get("code_columns", ["Your ticket code", "Ticket code"]))
    if not code_col:
        raise ValueError("The Google Forms export does not contain a recognized ticket-code column.")
    basket_columns, used_fallback = find_basket_columns(
        responses,
        form_config.get("basket_prefix", "Tickets for:"),
        reserved=[timestamp_col, name_col, code_col],
        ignore=form_config.get("ignore_columns", ["Email Address"]),
    )
    if used_fallback:
        print(
            f"No column headers begin with '{form_config.get('basket_prefix', 'Tickets for:')}'. "
            "Using every other column as a basket instead:"
        )
        for col in basket_columns:
            print(f"  - {col}")
        print(
            "If any of these are not real baskets (for example, an email-collection column), "
            "add the exact header text to \"ignore_columns\" under \"form\" in raffle_config.json, "
            "delete raffle_validation.xlsx, and rerun run_validate.bat."
        )

    responses = responses.copy()
    responses["_Source Row"] = range(2, len(responses) + 2)
    responses["_Normalized Code"] = responses[code_col].map(normalize_code)
    if timestamp_col:
        timestamp_values = responses[timestamp_col].astype(str).str.replace(
            r"\s+[A-Z]{2,5}$", "", regex=True
        )
        responses["_Parsed Timestamp"] = pd.to_datetime(
            timestamp_values, format="mixed", errors="coerce", utc=True
        )
    else:
        responses["_Parsed Timestamp"] = pd.NaT

    use_last = bool(config["validation"].get("use_last_response_per_code", True))
    responses["_Duplicate Status"] = ""
    active_indices: set[int] = set(responses.index)
    if use_last:
        for code, group in responses[responses["_Normalized Code"] != ""].groupby("_Normalized Code"):
            if len(group) <= 1:
                continue
            ordered = group.sort_values(["_Parsed Timestamp", "_Source Row"], na_position="first")
            keep_index = ordered.index[-1]
            for idx in ordered.index:
                responses.at[idx, "_Duplicate Status"] = "LATEST RESPONSE USED" if idx == keep_index else "OLDER DUPLICATE - NOT USED"
                if idx != keep_index:
                    active_indices.discard(idx)

    all_response_rows: list[dict[str, Any]] = []
    active_rows: list[dict[str, Any]] = []
    basket_totals = {col: {"submitted": 0, "eligible": 0, "entrants": 0} for col in basket_columns}

    for idx, source in responses.iterrows():
        code = source["_Normalized Code"]
        official = balance_by_code.get(code)
        official_name = clean_display(official.get("Participant Name")) if official else ""
        official_tickets = int(float(official.get("Official Tickets", 0))) if official else None
        submitted_name = clean_display(source.get(name_col)) if name_col else ""
        timestamp = clean_display(source.get(timestamp_col)) if timestamp_col else ""

        parsed_values: dict[str, int] = {}
        invalid_messages: list[str] = []
        allocated = 0
        for col in basket_columns:
            value, issue = parse_ticket_cell(source.get(col))
            parsed_values[col] = value
            allocated += value
            if issue:
                invalid_messages.append(f"{basket_name(col, form_config.get('basket_prefix', 'Tickets for:'))}: {issue}")

        is_active = idx in active_indices
        if not is_active:
            allocation_status = "DUPLICATE - OLDER RESPONSE"
            ready = "NO"
            difference: int | None = None
        elif not code:
            allocation_status = "MISSING CODE"
            ready = "NO"
            difference = None
        elif official is None:
            allocation_status = "UNKNOWN CODE"
            ready = "NO"
            difference = None
        elif invalid_messages:
            allocation_status = "INVALID TICKET VALUE"
            ready = "NO"
            difference = allocated - official_tickets
        else:
            difference = allocated - official_tickets
            if difference == 0:
                allocation_status = "EXACT"
                ready = "YES"
            elif difference < 0:
                allocation_status = "UNDER"
                ready = "YES" if config["validation"].get("allow_under_allocations", True) else "NO"
            else:
                allocation_status = "OVER"
                ready = "NO"

        if not submitted_name:
            name_check = "MISSING NAME"
        elif official_name and normalize_text(submitted_name) == normalize_text(official_name):
            name_check = "MATCH"
        elif official_name:
            name_check = "NAME DIFFERS"
        else:
            name_check = "NOT CHECKED"

        notes = "; ".join(invalid_messages)
        base = {
            "Source Row": int(source["_Source Row"]),
            "Timestamp": timestamp,
            "Submitted Name": submitted_name,
            "Ticket Code": code,
            "Official Name": official_name,
            "Official Tickets": official_tickets,
            "Allocated Tickets": allocated,
            "Difference": difference,
            "Allocation Status": allocation_status,
            "Name Check": name_check,
            "Ready for Draw": ready,
            "Duplicate Status": clean_display(source["_Duplicate Status"]),
            "Validation Notes": notes,
        }
        base.update(parsed_values)
        all_response_rows.append(base)
        if is_active:
            active_rows.append(base)
            for col, value in parsed_values.items():
                basket_totals[col]["submitted"] += value
                if ready == "YES" and value > 0:
                    basket_totals[col]["eligible"] += value
                    basket_totals[col]["entrants"] += 1

    active = pd.DataFrame(active_rows)
    all_responses = pd.DataFrame(all_response_rows)
    if active.empty:
        active = pd.DataFrame(columns=list(all_responses.columns))

    used_codes = set(active.loc[active["Ticket Code"].isin(balance_by_code), "Ticket Code"] if not active.empty else [])
    missing_responses = balances[~balances["Ticket Code"].isin(used_codes)][
        ["Participant Name", "Official Tickets", "Ticket Code"]
    ].copy()
    missing_responses["Status"] = "NO ACTIVE FORM RESPONSE"

    basket_summary_rows = []
    for col in basket_columns:
        basket_summary_rows.append(
            {
                "Basket Order": len(basket_summary_rows) + 1,
                "Basket": basket_name(col, form_config.get("basket_prefix", "Tickets for:")),
                "Validation Column": col,
                "Submitted Tickets": basket_totals[col]["submitted"],
                "Draw-Eligible Tickets": basket_totals[col]["eligible"],
                "Draw-Eligible Entrants": basket_totals[col]["entrants"],
            }
        )
    basket_summary = pd.DataFrame(basket_summary_rows)

    zero_baskets = [row["Basket"] for row in basket_summary_rows if row["Draw-Eligible Entrants"] == 0]
    if zero_baskets:
        carried = previous_basket_winners(paths.manual_basket_winners)
        manual_frame = pd.DataFrame(
            [{"Basket": name, "Manual Winner (optional)": carried.get(name, "")} for name in zero_baskets],
            columns=["Basket", "Manual Winner (optional)"],
        )
        write_workbook(paths.manual_basket_winners, [("Manual Basket Winners", manual_frame)])
        style_workbook(paths.manual_basket_winners, editable_columns={"Manual Basket Winners": ["Manual Winner (optional)"]})
        print(f"{len(zero_baskets)} basket(s) have zero draw-eligible tickets: {', '.join(zero_baskets)}")
        print(f"To hand one of these to someone anyway, open {paths.manual_basket_winners.name},")
        print("type their exact official name in \"Manual Winner (optional)\", save, and close it before running the draw.")
        print("Leave a row blank to skip that basket; it will not appear in the results or on the website.")
    elif paths.manual_basket_winners.exists():
        print(f"Every basket has at least one draw-eligible ticket. {paths.manual_basket_winners.name} is not needed and was left as-is.")

    counts = active["Allocation Status"].value_counts().to_dict() if not active.empty else {}
    blocking = active[active["Ready for Draw"] != "YES"] if not active.empty else active
    summary = pd.DataFrame(
        [
            ["Event", config["event"].get("name", "Clinic Fundraising Raffle")],
            ["Event Date", config["event"].get("date", "")],
            ["Validated At", now_text()],
            ["Form Export", str(paths.form_responses.relative_to(paths.root))],
            ["Active Responses", len(active)],
            ["Exact Allocations", counts.get("EXACT", 0)],
            ["Under Allocations", counts.get("UNDER", 0)],
            ["Over Allocations", counts.get("OVER", 0)],
            ["Invalid Ticket Values", counts.get("INVALID TICKET VALUE", 0)],
            ["Unknown Codes", counts.get("UNKNOWN CODE", 0)],
            ["Missing Codes", counts.get("MISSING CODE", 0)],
            ["Older Duplicate Responses Ignored", int((all_responses["Allocation Status"] == "DUPLICATE - OLDER RESPONSE").sum())],
            ["Participants Without an Active Response", len(missing_responses)],
            ["Blocking Responses", len(blocking)],
            ["Draw Status", "READY" if len(blocking) == 0 else "NOT READY - CORRECT THE FORM EXPORT AND REVALIDATE"],
            ["Rule", "Under-allocations remain eligible; over-allocations and invalid/unrecognized responses stop the draw."],
        ],
        columns=["Item", "Value"],
    )

    write_workbook(
        paths.validation,
        [
            ("Validation Summary", summary),
            ("Response Validation", active),
            ("Basket Totals", basket_summary),
            ("All Form Responses", all_responses),
            ("Missing Responses", missing_responses),
        ],
    )
    style_workbook(
        paths.validation,
        status_sheets=["Response Validation", "All Form Responses", "Missing Responses"],
    )

    print(f"Validation workbook created: {paths.validation}")
    print(f"Exact: {counts.get('EXACT', 0)} | Under: {counts.get('UNDER', 0)} | Over: {counts.get('OVER', 0)}")
    if len(blocking):
        print(f"DRAW NOT READY: {len(blocking)} blocking response(s). Review raffle_validation.xlsx.")
        print("Correct/re-export the Google Form, delete raffle_validation.xlsx, and rerun run_validate.bat.")
        return 2
    print("DRAW READY: no blocking responses were found.")
    return 0


def weighted_winner(entries: list[dict[str, Any]]) -> dict[str, Any]:
    total = sum(int(entry["Tickets"]) for entry in entries)
    pick = secrets.randbelow(total) + 1
    cumulative = 0
    for entry in entries:
        cumulative += int(entry["Tickets"])
        if pick <= cumulative:
            return entry
    return entries[-1]


def draw(config: dict[str, Any], paths: Paths) -> int:
    if not paths.validation.exists():
        raise FileNotFoundError("raffle_validation.xlsx does not exist. Run run_validate.bat first.")
    if paths.results.exists():
        raise FileExistsError(
            "raffle_results.xlsx already exists. It contains secret winners. Delete it only if you intentionally need to perform a new draw."
        )

    validation_summary = read_table(paths.validation, sheet_name="Validation Summary")
    responses = read_table(paths.validation, sheet_name="Response Validation")
    baskets = read_table(paths.validation, sheet_name="Basket Totals")
    require_columns(responses, ["Official Name", "Ticket Code", "Ready for Draw"], "Response Validation sheet")
    require_columns(baskets, ["Basket Order", "Basket", "Validation Column"], "Basket Totals sheet")

    if not paths.preparation.exists():
        raise FileNotFoundError("raffle_preparation.xlsx does not exist. Run run_prepare.bat first.")
    prep_balances = read_table(paths.preparation, sheet_name="Ticket Balances")
    require_columns(prep_balances, ["Participant Name", "Ticket Code"], "Ticket Balances sheet")
    official_roster: dict[str, tuple[str, str]] = {}
    for _, row in prep_balances.iterrows():
        official_name = clean_display(row["Participant Name"])
        if official_name and normalize_text(official_name) not in official_roster:
            official_roster[normalize_text(official_name)] = (official_name, normalize_code(row["Ticket Code"]))

    manual_winners: dict[str, str] = {}
    if paths.manual_basket_winners.exists():
        manual_frame = read_table(paths.manual_basket_winners, sheet_name="Manual Basket Winners")
        if {"Basket", "Manual Winner (optional)"}.issubset(manual_frame.columns):
            for _, row in manual_frame.iterrows():
                basket_name_value = clean_display(row["Basket"])
                winner_name = clean_display(row["Manual Winner (optional)"])
                if basket_name_value and winner_name:
                    manual_winners[basket_name_value] = winner_name

    blocking = responses[responses["Ready for Draw"].map(clean_display).str.upper() != "YES"]
    if not blocking.empty:
        print("Draw stopped. raffle_validation.xlsx still contains blocking responses:")
        display_cols = [col for col in ["Submitted Name", "Ticket Code", "Allocation Status", "Validation Notes"] if col in blocking.columns]
        print(blocking[display_cols].fillna("").to_string(index=False))
        print("Correct the form export, delete raffle_validation.xlsx, and rerun validation.")
        return 2

    allow_multiple = bool(config["draw"].get("allow_multiple_wins", True))
    previous_winners: set[str] = set()
    results_rows: list[dict[str, Any]] = []
    entries_rows: list[dict[str, Any]] = []
    drawn_at = now_text()

    for _, basket_row in baskets.sort_values("Basket Order").iterrows():
        column = clean_display(basket_row["Validation Column"])
        name = clean_display(basket_row["Basket"])
        if not column or column not in responses.columns:
            raise ValueError(f"Validation workbook is missing basket column: {column}")

        current_entries: list[dict[str, Any]] = []
        for _, response in responses.iterrows():
            tickets, issue = parse_ticket_cell(response.get(column))
            if issue:
                raise ValueError(f"Unexpected invalid value in approved validation workbook: {issue}")
            if tickets <= 0:
                continue
            participant = clean_display(response.get("Official Name"))
            code = normalize_code(response.get("Ticket Code"))
            if not allow_multiple and participant in previous_winners:
                continue
            entry = {
                "Basket Order": int(float(basket_row["Basket Order"])),
                "Basket": name,
                "Participant Name": participant,
                "Ticket Code": code,
                "Tickets": tickets,
            }
            current_entries.append(entry)
            entries_rows.append(entry)

        # A basket with zero eligible tickets is skipped unless the organizer
        # has typed a manual winner for it in manual_basket_winners.xlsx.
        if not current_entries:
            manual_name = manual_winners.get(name)
            if not manual_name:
                continue
            match = official_roster.get(normalize_text(manual_name))
            if not match:
                raise ValueError(
                    f"Manual winner \"{manual_name}\" for basket \"{name}\" in "
                    f"{paths.manual_basket_winners.name} does not match any official participant name "
                    "in raffle_preparation.xlsx. Fix the spelling to match exactly and rerun the draw."
                )
            official_name, official_code = match
            if not allow_multiple and official_name in previous_winners:
                raise ValueError(
                    f"Manual winner \"{official_name}\" for basket \"{name}\" has already won another basket, "
                    "and \"Multiple Wins Allowed\" is set to NO in raffle_config.json. "
                    "Choose someone else in manual_basket_winners.xlsx or allow multiple wins."
                )
            previous_winners.add(official_name)
            results_rows.append(
                {
                    "Basket Order": int(float(basket_row["Basket Order"])),
                    "Basket": name,
                    "Winner": official_name,
                    "Winner Ticket Code": official_code,
                    "Winner Tickets in Basket": 0,
                    "Total Tickets in Basket": 0,
                    "Entrants": 0,
                    "Manual Assignment": "YES",
                    "Drawn At": drawn_at,
                }
            )
            continue
        winner = weighted_winner(current_entries)
        previous_winners.add(winner["Participant Name"])
        pool = sum(entry["Tickets"] for entry in current_entries)
        results_rows.append(
            {
                "Basket Order": int(float(basket_row["Basket Order"])),
                "Basket": name,
                "Winner": winner["Participant Name"],
                "Winner Ticket Code": winner["Ticket Code"],
                "Winner Tickets in Basket": winner["Tickets"],
                "Total Tickets in Basket": pool,
                "Entrants": len(current_entries),
                "Manual Assignment": "NO",
                "Drawn At": drawn_at,
            }
        )

    manual_count = sum(1 for row in results_rows if row.get("Manual Assignment") == "YES")

    if not results_rows:
        raise ValueError("No positive ticket allocations were found. All zero values are ignored, so there is nothing to draw.")

    event_rows = pd.DataFrame(
        [
            ["Event", config["event"].get("name", "Clinic Fundraising Raffle")],
            ["Event Date", config["event"].get("date", "")],
            ["Results Created At", drawn_at],
            ["Source", "raffle_validation.xlsx"],
            ["Multiple Wins Allowed", "YES" if allow_multiple else "NO"],
            ["Manually Assigned Baskets", manual_count],
            ["Website File", "draw.html"],
            ["Confidentiality", "SECRET UNTIL THE LIVE RAFFLE"],
        ],
        columns=["Item", "Value"],
    )
    results = pd.DataFrame(results_rows).sort_values("Basket Order")
    entries = pd.DataFrame(entries_rows).sort_values(["Basket Order", "Participant Name"])
    write_workbook(paths.results, [("Event", event_rows), ("Results", results), ("Entries", entries)])
    style_workbook(paths.results)

    # website data: one basket per result row, entrants aggregated per participant
    web_baskets: list[dict[str, Any]] = []
    for row in results_rows:
        totals: dict[str, int] = {}
        for entry in entries_rows:
            if entry["Basket Order"] == row["Basket Order"]:
                totals[entry["Participant Name"]] = totals.get(entry["Participant Name"], 0) + entry["Tickets"]
        web_baskets.append(
            {
                "name": row["Basket"],
                "pool": row["Total Tickets in Basket"],
                "winner": row["Winner"],
                "winnerTickets": totals.get(row["Winner"], row["Winner Tickets in Basket"]),
                "entrants": [{"n": name, "t": tickets} for name, tickets in sorted(totals.items())],
                "manual": row.get("Manual Assignment") == "YES",
            }
        )
    payload = {
        "event": config["event"].get("name", "Clinic Fundraising Raffle"),
        "date": config["event"].get("date", ""),
        "drawnAt": drawn_at,
        "baskets": web_baskets,
    }
    with paths.website_data.open("w", encoding="utf-8") as handle:
        handle.write("const DATA = " + json.dumps(payload, indent=2) + ";\n")

    print(f"SECRET results created: {paths.results}")
    print(f"SECRET website data created: {paths.website_data}")
    print(f"Baskets drawn: {len(results):,}")
    if manual_count:
        print(f"Of those, {manual_count} basket(s) were manually assigned (zero draw-eligible tickets).")
    print("Do not open or share raffle_results.xlsx or draw_data.js before the live raffle.")
    print("At the event, run run_website.bat to present the winners one basket at a time.")
    return 0


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Prepare, validate, and draw a clinic e-raffle.")
    parser.add_argument("command", choices=["prepare", "validate", "draw"])
    parser.add_argument("--config", default="raffle_config.json")
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    config_path = Path(args.config).resolve()
    try:
        config = load_config(config_path)
        paths = build_paths(config_path, config, args.command)
        if args.command == "prepare":
            return prepare(config, paths)
        if args.command == "validate":
            return validate(config, paths)
        return draw(config, paths)
    except Exception as exc:
        print(f"ERROR: {exc}", file=sys.stderr)
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
